from __future__ import annotations

import csv
import json
import os
import time
from typing import Any, Dict, List, Optional
import urllib.request
from urllib.parse import unquote

import itertools
import numpy as np
import pandas as pd

from fastapi.responses import Response
from fastapi import FastAPI, Header, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
from openpyxl import load_workbook


# Env

APP_SECRET = os.getenv("AGENT_REG_SECRET", "dev-secret")
ORCHESTRATOR_BASE_URL = os.getenv("ORCHESTRATOR_BASE_URL", "http://backend:8000/api/v1")
NODE_NAME = os.getenv("NODE_NAME", "Hospital Agent")
HOSPITAL_ORG = os.getenv("HOSPITAL_ORG", "HospitalA")
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "http://hospital_a_agent:9001")
DATA_DIR = os.getenv("DATA_DIR", "/data")

MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(200 * 1024 * 1024)))  # Μέγιστο μέγεθος upload για προστασία
OUTLIER_Z = float(os.getenv("OUTLIER_Z", "3.0"))
MIN_ROWS_THRESHOLD = int(os.getenv("MIN_ROWS_THRESHOLD", "50")) # Privacy threshold: κάτω από αυτό -> suppression (δεν επιστρέφει aggregates) (fl-job)


# Consent filtering

CONSENT_FILTER_ENABLED = str(os.getenv("CONSENT_FILTER_ENABLED", "0")).strip().lower() in ("1", "true", "yes", "on")
PATIENT_ID_COLUMN = os.getenv("PATIENT_ID_COLUMN", "patient_id")
PATIENT_PORTAL_SECRET = os.getenv("PATIENT_PORTAL_SECRET", "")
ENRICHED_METRICS_ENABLED = str(os.getenv("ENRICHED_METRICS_ENABLED", "1")).strip().lower() in ("1", "true", "yes", "on")
ENRICHED_SAMPLE_CAP = int(os.getenv("ENRICHED_SAMPLE_CAP", "20000"))

# Δημιουργία FastAPI
app = FastAPI(title="BCFL Hospital Agent")


# Models

class ValidateRequest(BaseModel):
    local_uri: str = Field(..., min_length=1)
    schema_id: str = Field(..., min_length=1)


class TrainRoundRequest(BaseModel):
    job_id: str
    round: int
    dataset_id: str
    local_uri: str
    schema_id: str
    features: List[str] = Field(default_factory=list)
    label: Optional[str] = None
    stratify_by: Optional[str] = None


class PatientConsentLinkRequest(BaseModel):
    dataset_id: str = Field(..., min_length=1)
    patient_id: str = Field(..., min_length=1)


# Helpers

def _check_secret(x_agent_secret: Optional[str]) -> None:
    if x_agent_secret != APP_SECRET:
        raise HTTPException(status_code=403, detail="Invalid agent secret")


def _ensure_data_dir() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)


def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        s = str(x).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _read_csv_head(path: str, max_rows: int = 5000) -> Dict[str, Any]:
    if not os.path.exists(path):
        return {"ok": False, "error": f"File not found: {path}"}

    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        n = 0
        for _ in reader:
            n += 1
            if n >= max_rows:
                break
    return {"ok": True, "columns": cols, "row_count": n}


def _read_xlsx_head(path: str, max_rows: int = 5000) -> Dict[str, Any]:
    if not os.path.exists(path):
        return {"ok": False, "error": f"File not found: {path}"}

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {"ok": False, "error": "Empty Excel file"}

    cols = [str(c).strip() for c in header_row if c is not None and str(c).strip() != ""]

    n = 0
    for _ in ws.iter_rows(min_row=2, max_row=1 + max_rows, values_only=True):
        n += 1

    return {"ok": True, "columns": cols, "row_count": n}


def _detect_columns(path: str) -> Dict[str, Any]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_csv_head(path)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _read_xlsx_head(path)
    return {"ok": False, "error": f"Unsupported file type: {ext}. Allowed: .csv, .xlsx"}


def _http_post(url: str, payload: dict, timeout: int = 8) -> dict:
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8")
        return json.loads(raw) if raw else {}


def _normalize_gender(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s == "":
        return None

    # Αποδεκτές τιμές/συνώνυμα
    if s in ("m", "male", "man", "άνδρας"):
        return "male"
    if s in ("f", "female", "woman", "γυναίκα", "θήλυ", "θηλυκό"):
        return "female"
    if s in ("other", "nonbinary", "non-binary", "nb", "άλλο"):
        return "other"

    return None

def _compute_gender_stratified_metrics(
    df: pd.DataFrame,
    gender_col: str,
    features: List[str],
    outlier_z: float,
    min_rows_threshold: int,
) -> Dict[str, Any]:
    if not gender_col or gender_col not in df.columns:
        return {}

    work = df.copy()
    work[gender_col] = work[gender_col].apply(_normalize_gender)
    work = work.dropna(subset=[gender_col])

    out: Dict[str, Any] = {}

    for gval, gdf in work.groupby(gender_col):
        group_size = int(len(gdf))

        # Privacy threshold ανά φύλο
        if group_size < int(min_rows_threshold):
            out[str(gval)] = {
                "suppressed": True,
                "row_count": group_size,
                "reason": f"group < {min_rows_threshold}",
            }
            continue

        fm = _compute_feature_metrics_df(gdf, features, outlier_z)

        out[str(gval)] = {
            "suppressed": False,
            "row_count": group_size,
            "feature_metrics": fm,
        }

    return out

# Consent

_cons_cache: Dict[str, Dict[str, bool]] = {}


def _orchestrator_has_consent(dataset_id: str, patient_id: str) -> bool:

    try:
        url = f"{ORCHESTRATOR_BASE_URL.rstrip('/')}/consents/has"
        payload = {"dataset_id": dataset_id, "patient_id": patient_id}
        data = _http_post(url, payload, timeout=8)
        return bool(data.get("has_consent"))
    except Exception:
        return False


def _has_consent_cached(dataset_id: str, patient_id: str) -> bool:

    ds = (dataset_id or "").strip()
    pid = (patient_id or "").strip()
    if not ds or not pid:
        return False

    ds_cache = _cons_cache.setdefault(ds, {})
    if pid in ds_cache:
        return ds_cache[pid]

    ok = _orchestrator_has_consent(ds, pid)
    ds_cache[pid] = ok
    return ok


# Training metrics

def _iqr_stats(series: pd.Series) -> Dict[str, float]:
    q1 = float(series.quantile(0.25)) if series.notna().any() else 0.0
    q3 = float(series.quantile(0.75)) if series.notna().any() else 0.0
    return {"q1": q1, "q3": q3, "iqr": float(q3 - q1)}


def _outlier_rate_zscore(series: pd.Series, z: float = 3.0) -> float:
    s = series.dropna()
    if len(s) < 5:
        return 0.0
    std = float(s.std(ddof=0))
    if std == 0.0:
        return 0.0
    mean = float(s.mean())
    zz = ((s - mean) / std).abs()
    return float((zz > z).mean())

def _compute_enriched_feature_metrics_df(df: pd.DataFrame, features: List[str]) -> Dict[str, Any]:
    feature_metrics: Dict[str, Dict[str, Any]] = {}
    variances: Dict[str, float] = {}

    # Κρατάμε μόνο features που υπάρχουν στο df
    present_features = [f for f in (features or []) if f in df.columns]

    for feat in present_features:
        col = df[feat]
        missing_rate = float(col.isna().mean())

        num = pd.to_numeric(col, errors="coerce")
        non_na = num.dropna()

        unique_values = int(col.nunique(dropna=True))
        is_constant = bool(unique_values <= 1)

        mean = float(non_na.mean()) if len(non_na) else 0.0
        std = float(non_na.std(ddof=0)) if len(non_na) else 0.0
        vmin = float(non_na.min()) if len(non_na) else 0.0
        vmax = float(non_na.max()) if len(non_na) else 0.0
        median = float(non_na.median()) if len(non_na) else 0.0

        iqr = _iqr_stats(non_na) if len(non_na) else {"q1": 0.0, "q3": 0.0, "iqr": 0.0}
        outlier_rate = _outlier_rate_zscore(non_na, z=3.0) if len(non_na) else 0.0

        var = float(non_na.var(ddof=0)) if len(non_na) else 0.0
        variances[feat] = var

        feature_metrics[feat] = {
            "mean": mean,
            "std": std,
            "min": vmin,
            "max": vmax,
            "median": median,
            "q1": float(iqr["q1"]),
            "q3": float(iqr["q3"]),
            "iqr": float(iqr["iqr"]),
            "missing_rate": missing_rate,
            "outlier_rate": float(outlier_rate),
            "unique_values": unique_values,
            "is_constant": is_constant,
            "variance": var,
        }

    total_var = float(sum(max(v, 0.0) for v in variances.values()))
    normalized_importance = {
        k: (float(max(v, 0.0)) / total_var if total_var > 0 else 0.0)
        for k, v in variances.items()
    }

    # Pearson correlation matrix
    corr_matrix: Dict[str, Dict[str, float]] = {}
    if len(present_features) >= 2:
        numeric_df = df[present_features].apply(pd.to_numeric, errors="coerce")
        numeric_df = numeric_df.dropna(axis=1, how="all")
        if numeric_df.shape[1] >= 2:
            corr = numeric_df.corr(method="pearson")
            corr_matrix = {c: {r: float(corr.loc[r, c]) for r in corr.index} for c in corr.columns}

    return {
        "feature_metrics": feature_metrics,
        "normalized_importance": normalized_importance,
        "correlation_matrix": corr_matrix,
        "present_features": present_features,  # χρήσιμο debug
    }

def _compute_feature_means_csv(
    path: str,
    features: List[str],
    dataset_id: str,
    max_rows: int = 200000,
) -> Dict[str, Any]:

    if not os.path.exists(path):
        return {"ok": False, "error": f"File not found: {path}"}

    sums: Dict[str, float] = {f: 0.0 for f in features}
    counts: Dict[str, int] = {f: 0 for f in features}

    rows_total = 0
    rows_used = 0
    rows_skipped_no_consent = 0

    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        header_cols = set(reader.fieldnames or [])

        for row in reader:
            rows_total += 1

            if CONSENT_FILTER_ENABLED:
                pid = (row.get(PATIENT_ID_COLUMN) or "").strip()
                if not pid or not _has_consent_cached(dataset_id, pid):
                    rows_skipped_no_consent += 1
                    if rows_total >= max_rows:
                        break
                    continue

            rows_used += 1

            for feat in features:
                if feat not in header_cols:
                    continue

                v = _safe_float(row.get(feat))
                if v is not None:
                    sums[feat] += v
                    counts[feat] += 1

            if rows_total >= max_rows:
                break

    means: Dict[str, float] = {}
    missing_features: List[str] = []

    # αν count==0 -> δεν το βάζουμε στο update
    for feat in features:
        if counts[feat] > 0:
            means[feat] = sums[feat] / float(counts[feat])
        else:
            missing_features.append(feat)

    return {
        "ok": True,
        "row_count": rows_total,
        "update": means,
        "metrics": {
            "rows_processed": rows_total,
            "rows_used": rows_used,
            "rows_skipped_no_consent": rows_skipped_no_consent,
            "consent_filter_enabled": CONSENT_FILTER_ENABLED,
            "patient_id_column": PATIENT_ID_COLUMN if CONSENT_FILTER_ENABLED else None,
            "missing_features": missing_features,
        },
    }

def _compute_feature_means_xlsx(
    path: str,
    features: List[str],
    dataset_id: str,
    max_rows: int = 200000,
) -> Dict[str, Any]:

    if not os.path.exists(path):
        return {"ok": False, "error": f"File not found: {path}"}

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {"ok": False, "error": "Empty Excel file"}

    cols = [str(c).strip() if c is not None else "" for c in header_row]
    col_index = {name: idx for idx, name in enumerate(cols) if name}

    pid_idx = col_index.get(PATIENT_ID_COLUMN) if CONSENT_FILTER_ENABLED else None

    sums: Dict[str, float] = {f: 0.0 for f in features}
    counts: Dict[str, int] = {f: 0 for f in features}

    rows_total = 0
    rows_used = 0
    rows_skipped_no_consent = 0

    for row in ws.iter_rows(min_row=2, max_row=1 + max_rows, values_only=True):
        rows_total += 1

        if CONSENT_FILTER_ENABLED:
            pid = ""
            if pid_idx is not None and pid_idx < len(row):
                pid = str(row[pid_idx] or "").strip()
            if not pid or not _has_consent_cached(dataset_id, pid):
                rows_skipped_no_consent += 1
                continue

        rows_used += 1

        for feat in features:
            idx = col_index.get(feat)
            if idx is None or idx >= len(row):
                continue
            v = _safe_float(row[idx])
            if v is not None:
                sums[feat] += v
                counts[feat] += 1

    means: Dict[str, float] = {}
    missing_features: List[str] = []

    for feat in features:
        if counts[feat] > 0:
            means[feat] = sums[feat] / float(counts[feat])
        else:
            missing_features.append(feat)

    return {
        "ok": True,
        "row_count": rows_total,
        "update": means,
        "metrics": {
            "rows_processed": rows_total,
            "rows_used": rows_used,
            "rows_skipped_no_consent": rows_skipped_no_consent,
            "consent_filter_enabled": CONSENT_FILTER_ENABLED,
            "patient_id_column": PATIENT_ID_COLUMN if CONSENT_FILTER_ENABLED else None,
            "missing_features": missing_features,
        },
    }

def _compute_feature_means(
        path: str, features: List[str],
        dataset_id: str,
        max_rows: int = 200000)\
            -> Dict[str, Any]:

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _compute_feature_means_csv(path, features, dataset_id=dataset_id, max_rows=max_rows)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _compute_feature_means_xlsx(path, features, dataset_id=dataset_id, max_rows=max_rows)
    return {"ok": False, "error": f"Unsupported file type for training: {ext}"}


def _read_df_for_metrics(path: str, columns: List[str], max_rows: int) -> pd.DataFrame:

    ext = os.path.splitext(path)[1].lower()
    usecols = [c for c in (columns or []) if c]

    if ext == ".csv":
        return pd.read_csv(path, usecols=usecols, nrows=max_rows, low_memory=False)
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return pd.read_excel(path, usecols=usecols, nrows=max_rows)
    raise ValueError(f"Unsupported file type for metrics: {ext}")


def _compute_feature_metrics_df(df: pd.DataFrame, features: List[str], outlier_z: float) -> Dict[str, Any]:

    res: Dict[str, Any] = {}

    for f in features:
        if f not in df.columns:
            continue

        s_raw = df[f]
        total = int(len(s_raw))
        missing = int(s_raw.isna().sum())

        s_num = pd.to_numeric(s_raw, errors="coerce")
        non_na = s_num.dropna()

        if len(non_na) > 0:
            # numeric case
            x = non_na.to_numpy(dtype=float)
            n = int(x.size)

            mean = float(np.mean(x))
            std = float(np.std(x, ddof=0))
            vmin = float(np.min(x))
            vmax = float(np.max(x))

            q1 = float(np.quantile(x, 0.25))
            med = float(np.quantile(x, 0.50))
            q3 = float(np.quantile(x, 0.75))
            iqr = float(q3 - q1)

            outliers = 0
            if std > 0:
                z = np.abs((x - mean) / std)
                outliers = int((z > outlier_z).sum())

            missing_rate = float(missing / max(total, 1))
            outlier_rate = float(outliers / max(n, 1))

            uniq = int(pd.Series(x).nunique(dropna=True))
            is_constant = bool(uniq <= 1)

            res[f] = {
                "type": "numeric",
                "total": total,
                "missing": missing,
                "n": n,
                "mean": mean,
                "std": std,
                "min": vmin,
                "max": vmax,
                "q1": q1,
                "median": med,
                "q3": q3,
                "iqr": iqr,
                "missing_rate": missing_rate,
                "outliers": outliers,
                "outlier_rate": outlier_rate,
                "unique": uniq,
                "is_constant": is_constant,
            }
        else:
            non_na_raw = s_raw.dropna()
            uniq = int(non_na_raw.nunique(dropna=True))
            denom = max(int(len(non_na_raw)), 1)

            res[f] = {
                "type": "categorical",
                "total": total,
                "missing": missing,
                "missing_rate": float(missing / max(total, 1)),
                "unique": uniq,
                "unique_ratio": float(uniq / denom) if len(non_na_raw) else 0.0,
                "is_constant": bool(uniq <= 1),
            }

    return res

#Ομαδοποιεί το dataset με βάση τη στήλη stratify_by και για κάθε ομάδα, υπολογίζει feature metrics
def _compute_stratified_metrics(df: pd.DataFrame, stratify_by: str, features: List[str], outlier_z: float) -> Dict[str, Any]:
    if not stratify_by or stratify_by not in df.columns:
        return {}

    out: Dict[str, Any] = {}

    for gval, gdf in df.dropna(subset=[stratify_by]).groupby(df[stratify_by].astype(str)):
        out[str(gval)] = _compute_feature_metrics_df(gdf, features, outlier_z)

    return out

def _compute_sufficient_stats_df(df: pd.DataFrame, features: List[str]) -> Dict[str, Any]:
    present = [f for f in (features or []) if f in df.columns]
    if not present:
        return {"present_features": [], "feature_sums": {}, "pair_sums": {}}

    numeric_df = df[present].apply(pd.to_numeric, errors="coerce")

    feature_sums: Dict[str, dict] = {}
    for f in present:
        s = numeric_df[f]
        total_rows = int(len(s))
        n = int(s.notna().sum())
        missing = int(total_rows - n)

        if n > 0:
            x = s.dropna().to_numpy(dtype=float)
            feature_sums[f] = {
                "total_rows": total_rows,
                "n": n,
                "missing": missing,
                "sum": float(np.sum(x)),
                "sumsq": float(np.sum(x * x)),
                "min": float(np.min(x)),
                "max": float(np.max(x)),
            }
        else:
            feature_sums[f] = {
                "total_rows": total_rows,
                "n": 0,
                "missing": missing,
                "sum": 0.0,
                "sumsq": 0.0,
                "min": None,
                "max": None,
            }

    pair_sums: Dict[str, dict] = {}
    for a, b in itertools.combinations(present, 2):
        xa = numeric_df[a]
        xb = numeric_df[b]
        mask = xa.notna() & xb.notna()
        n = int(mask.sum())
        key = f"{a}||{b}"
        if n > 0:
            A = xa[mask].to_numpy(dtype=float)
            B = xb[mask].to_numpy(dtype=float)
            pair_sums[key] = {
                "n": n,
                "sum_x": float(np.sum(A)),
                "sum_y": float(np.sum(B)),
                "sum_x2": float(np.sum(A * A)),
                "sum_y2": float(np.sum(B * B)),
                "sum_xy": float(np.sum(A * B)),
            }
        else:
            pair_sums[key] = {"n": 0, "sum_x": 0.0, "sum_y": 0.0, "sum_x2": 0.0, "sum_y2": 0.0, "sum_xy": 0.0}

    return {
        "present_features": present,
        "feature_sums": feature_sums,
        "pair_sums": pair_sums,
    }

# Routes

@app.get("/health")
def health():
    return {
        "status": "ok",
        "org": HOSPITAL_ORG,
        "name": NODE_NAME
    }


@app.post("/upload")
async def upload(file: UploadFile = File(...), x_agent_secret: Optional[str] = Header(default=None)):
    _check_secret(x_agent_secret)
    _ensure_data_dir()

    contents = await file.read()
    if contents is None:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large. Max is {MAX_UPLOAD_BYTES} bytes")

    filename = (file.filename or "upload").strip()
    ext = os.path.splitext(filename)[1].lower() # έλεγχος τύπου αρχείου
    if ext not in (".csv", ".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {ext}. Allowed: CSV, XLSX")

    safe_name = filename.replace(" ", "_") # αντικατάσταση κενών αν χρειάζεται στο όνομα)
    local_uri = os.path.join(DATA_DIR, f"{int(time.time())}_{safe_name}") # με timestamp ώστε να μην μπερδεύονται


    with open(local_uri, "wb") as f:
        f.write(contents)  # Αποθηκεύει το αρχείο στο data dir ως bytes

    info = _detect_columns(local_uri)
    if not info.get("ok"):
        raise HTTPException(status_code=400, detail=info.get("error", "Failed to read columns"))

    return {
        "ok": True,
        "local_uri": local_uri,
        "columns": info.get("columns", []),
        "row_count": info.get("row_count", 0),
        "filename": filename,
    }


@app.post("/validate")
def validate(req: ValidateRequest, x_agent_secret: Optional[str] = Header(default=None)):
    _check_secret(x_agent_secret)
    info = _detect_columns(req.local_uri)
    info["schema_id"] = req.schema_id
    return info


@app.post("/train_round") #Δεν επιστρέφει raw data - Επιστρέφει "updates" (aggregates) + optional metrics.
def train_round(req: TrainRoundRequest, x_agent_secret: Optional[str] = Header(default=None)):
    _check_secret(x_agent_secret)

    features = req.features or []
    if not features:
        return {
            "ok": False,
            "error": "Provide features list",
            "row_count": 0,
            "update": {},
            "suppressed": False,
            "min_row_threshold": MIN_ROWS_THRESHOLD,
            "metrics": {
                "feature_metrics": {},
                "stratified_metrics": {},
                "privacy": {"min_row_threshold": MIN_ROWS_THRESHOLD},
            },
        }


    base = _compute_feature_means(req.local_uri, features=features, dataset_id=req.dataset_id)
    if not base.get("ok"):
        return {
            "ok": False,
            "error": base.get("error", "compute failed"),
            "row_count": int(base.get("row_count") or 0),
            "update": {},
            "feature_metrics": {},
            "stratified_metrics": {},
            "suppressed": False,
            "min_row_threshold": MIN_ROWS_THRESHOLD,
            "metrics": base.get("metrics") or {},
        }

    row_count = int(base.get("row_count") or 0)
    effective_row_count = int((base.get("metrics") or {}).get("rows_used") or 0)

    if effective_row_count < MIN_ROWS_THRESHOLD:
        return {
            "ok": True,
            "row_count": row_count,
            "update": {},
            "suppressed": True,
            "min_row_threshold": MIN_ROWS_THRESHOLD,
            "metrics": {
                **(base.get("metrics") or {}),
                "privacy": {"min_row_threshold": MIN_ROWS_THRESHOLD, "suppressed": True},
                "effective_row_count": effective_row_count,
                "sufficient_stats": {},
            },
        }

    sufficient_stats = {"present_features": [], "feature_sums": {}, "pair_sums": {}}
    correlation_matrix = {}
    normalized_importance = {}
    feature_metrics = {}
    stratified_metrics = {}

    if not CONSENT_FILTER_ENABLED:
        try:
            cols_needed = list(features)
            if req.stratify_by:
                cols_needed.append(req.stratify_by)

            df = _read_df_for_metrics(req.local_uri, cols_needed, max_rows=200000)
            sufficient_stats = _compute_sufficient_stats_df(df, features)

            if ENRICHED_METRICS_ENABLED:
                df_enriched = df
                if ENRICHED_SAMPLE_CAP and len(df) > ENRICHED_SAMPLE_CAP:
                    df_enriched = df.sample(n=ENRICHED_SAMPLE_CAP, random_state=42)

                enriched = _compute_enriched_feature_metrics_df(df_enriched, features)
                feature_metrics = enriched.get("feature_metrics", {})
                correlation_matrix = enriched.get("correlation_matrix", {})
                normalized_importance = enriched.get("normalized_importance", {})
            else:
                feature_metrics = _compute_feature_metrics_df(df, features, OUTLIER_Z)

            if req.stratify_by and req.stratify_by.strip().lower() == "gender":
                stratified_metrics = _compute_gender_stratified_metrics(
                    df=df,
                    gender_col=req.stratify_by,
                    features=features,
                    outlier_z=OUTLIER_Z,
                    min_rows_threshold=MIN_ROWS_THRESHOLD,
                )
        except Exception:
            pass

    return {
        "ok": True,
        "row_count": row_count,
        "update": base.get("update") or {},
        "suppressed": False,
        "min_row_threshold": MIN_ROWS_THRESHOLD,
        "metrics": {
            **(base.get("metrics") or {}),
            "privacy": {"min_row_threshold": MIN_ROWS_THRESHOLD, "suppressed": False},
            "effective_row_count": effective_row_count,
            "outlier_z": OUTLIER_Z,
            "requested_features": features,
            "missing_features": (base.get("metrics") or {}).get("missing_features", []),

            "feature_metrics": feature_metrics,
            "stratified_metrics": stratified_metrics,
            "sufficient_stats": sufficient_stats,
            "correlation_matrix": correlation_matrix,
            "normalized_importance": normalized_importance,
        },
    }


@app.post("/patient/consent-link")
def patient_consent_link(req: PatientConsentLinkRequest, x_agent_secret: Optional[str] = Header(default=None)):
    _check_secret(x_agent_secret)

    if not PATIENT_PORTAL_SECRET:
        raise HTTPException(status_code=500, detail="PATIENT_PORTAL_SECRET is not set on agent")

    token_payload = {
        "dataset_id": req.dataset_id,
        "patient_id": req.patient_id,
        "org": HOSPITAL_ORG,
        "ts": int(time.time()),
    }

    # το JSON έχει πάντα την ίδια μορφή -άρα το hash θα είναι ίδιο για τα ίδια δεδομένα (deterministic serialization.)
    raw = json.dumps(token_payload, separators=(",", ":"), sort_keys=True).encode("utf-8")

    import hashlib
    sig = hashlib.sha256((PATIENT_PORTAL_SECRET.encode("utf-8") + raw)).hexdigest() # Μπορεί να αλλάξει αργότερα με personalized authenticator

    return {"ok": True, "token": sig, "payload": token_payload}



@app.get("/download")
def download_file(
    local_uri: str,
    x_agent_secret: Optional[str] = Header(default=None),
):

    _check_secret(x_agent_secret)

    path = (local_uri or "").strip()
    path = unquote(path)

    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")

    if not os.path.abspath(path).startswith(os.path.abspath(DATA_DIR)):
        raise HTTPException(status_code=403, detail="Access outside data dir not allowed")

    with open(path, "rb") as f:
        data = f.read()

    return Response(content=data, media_type="application/octet-stream")

@app.on_event("startup")
def startup_register():
    try:
        url = f"{ORCHESTRATOR_BASE_URL.rstrip('/')}/nodes/register?secret={APP_SECRET}"
        payload = {"org": HOSPITAL_ORG, "base_url": PUBLIC_BASE_URL, "name": NODE_NAME}
        _http_post(url, payload, timeout=8)
    except Exception:
        pass

